"""Safe XLSX workbook operations for CODEGA AI.

This module keeps Excel handling local-first and path-confined. Heavy
dependencies are imported lazily so the desktop app can boot even when the
spreadsheet stack is not installed yet.
"""

from __future__ import annotations

import re
import uuid
import zipfile
from pathlib import Path
from typing import Any

from codegaai.config import OUTPUTS_DIR, TEMP_DIR

MAX_XLSX_BYTES = 25 * 1024 * 1024
XLSX_INPUT_DIR = TEMP_DIR / "xlsx"
XLSX_OUTPUT_DIR = OUTPUTS_DIR / "xlsx"
ALLOWED_XLSX_ROOTS = (XLSX_INPUT_DIR, XLSX_OUTPUT_DIR)


class XlsxSkillError(ValueError):
    """User-safe XLSX skill error."""


class XlsxDependencyError(RuntimeError):
    """Raised when pandas/openpyxl are not available."""


def _deps() -> tuple[Any, Any]:
    try:
        import pandas as pd  # type: ignore[import-not-found]
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:
        raise XlsxDependencyError(
            "XLSX skill requires pandas and openpyxl. Install requirements.txt."
        ) from exc
    return pd, openpyxl


def _ensure_dirs() -> None:
    XLSX_INPUT_DIR.mkdir(parents=True, exist_ok=True)
    XLSX_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _is_subpath(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def resolve_allowed_xlsx_path(path: str | Path) -> Path:
    _ensure_dirs()
    candidate = Path(path).expanduser().resolve()
    if not any(_is_subpath(candidate, root) for root in ALLOWED_XLSX_ROOTS):
        raise XlsxSkillError("Path is outside allowed XLSX workspace roots.")
    if candidate.suffix.lower() != ".xlsx":
        raise XlsxSkillError("Only .xlsx workbooks are supported.")
    return candidate


def safe_xlsx_filename(name: str, fallback: str = "workbook.xlsx") -> str:
    raw = Path(name or fallback).name
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(raw).stem).strip("._")
    if not stem:
        stem = Path(fallback).stem
    return f"{stem[:80]}.xlsx"


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]+", "_", str(name or "Sheet")).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:31]


def save_uploaded_workbook(filename: str, content: bytes) -> Path:
    _ensure_dirs()
    if len(content) > MAX_XLSX_BYTES:
        raise XlsxSkillError("XLSX file is too large for the local workspace.")
    safe_name = safe_xlsx_filename(filename)
    target = XLSX_INPUT_DIR / f"{uuid.uuid4().hex[:10]}-{safe_name}"
    probe = _bytes_path_probe(content)
    if not zipfile.is_zipfile(probe):
        raise XlsxSkillError("Invalid .xlsx file structure.")
    with zipfile.ZipFile(_bytes_path_probe(content)) as archive:
        names = set(archive.namelist())
        if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
            raise XlsxSkillError("Invalid .xlsx workbook structure.")
    target.write_bytes(content)
    return resolve_allowed_xlsx_path(target)


def _bytes_path_probe(content: bytes) -> Any:
    import io

    return io.BytesIO(content)


def inspect_workbook(path: str | Path, preview_rows: int = 20) -> dict[str, Any]:
    pd, openpyxl = _deps()
    workbook_path = resolve_allowed_xlsx_path(path)
    if not workbook_path.exists():
        raise XlsxSkillError("Workbook was not found.")

    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
    excel = pd.ExcelFile(workbook_path, engine="openpyxl")
    try:
        sheets: list[dict[str, Any]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            df = pd.read_excel(excel, sheet_name=sheet_name, nrows=max(1, preview_rows))
            sheets.append(
                {
                    "name": sheet_name,
                    "rows": int(ws.max_row or 0),
                    "columns": int(ws.max_column or 0),
                    "merged_cells": [str(rng) for rng in ws.merged_cells.ranges],
                    "headers": [str(col) for col in df.columns.tolist()],
                    "dtypes": {str(k): str(v) for k, v in df.dtypes.items()},
                    "non_null": {str(k): int(v) for k, v in df.count().items()},
                    "preview": _records(df),
                }
            )
        props = wb.properties
        metadata = {
            "creator": props.creator,
            "last_modified_by": props.lastModifiedBy,
            "created": props.created.isoformat() if props.created else None,
            "modified": props.modified.isoformat() if props.modified else None,
        }
    finally:
        excel.close()
        wb.close()
    return {
        "filename": workbook_path.name,
        "path": str(workbook_path),
        "sheet_count": len(sheets),
        "sheets": sheets,
        "metadata": metadata,
    }


def _records(df: Any) -> list[dict[str, Any]]:
    records = df.where(df.notnull(), None).to_dict(orient="records")
    return [{str(k): _json_value(v) for k, v in row.items()} for row in records]


def _json_value(value: Any) -> Any:
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def apply_workbook_operations(
    source_path: str | Path,
    operations: list[dict[str, Any]],
    output_name: str | None = None,
) -> dict[str, Any]:
    _, openpyxl = _deps()
    workbook_path = resolve_allowed_xlsx_path(source_path)
    if not workbook_path.exists():
        raise XlsxSkillError("Workbook was not found.")
    if len(operations) > 100:
        raise XlsxSkillError("Too many workbook operations in one request.")

    wb = openpyxl.load_workbook(workbook_path)
    applied: list[str] = []
    for op in operations:
        action = str(op.get("action") or "").strip().lower()
        if action == "create_sheet":
            applied.append(_op_create_sheet(wb, op))
        elif action == "update_cell":
            applied.append(_op_update_cell(wb, op))
        elif action == "insert_row":
            applied.append(_op_insert_row(wb, op))
        elif action == "delete_row":
            applied.append(_op_delete_row(wb, op))
        elif action == "insert_column":
            applied.append(_op_insert_column(wb, op))
        elif action == "delete_column":
            applied.append(_op_delete_column(wb, op))
        elif action == "format_range":
            applied.append(_op_format_range(wb, op, openpyxl))
        elif action == "add_chart":
            applied.append(_op_add_chart(wb, op, openpyxl))
        else:
            raise XlsxSkillError(f"Unsupported XLSX operation: {action or 'empty'}")

    safe_output = safe_xlsx_filename(output_name or f"{workbook_path.stem}-edited.xlsx")
    output_path = resolve_allowed_xlsx_path(XLSX_OUTPUT_DIR / safe_output)
    wb.save(output_path)
    wb.close()
    return {
        "ok": True,
        "output_path": str(output_path),
        "output_url": f"/outputs/xlsx/{output_path.name}",
        "applied": applied,
        "summary": inspect_workbook(output_path, preview_rows=10),
    }


def _sheet(wb: Any, name: str) -> Any:
    sheet_name = safe_sheet_name(name)
    if sheet_name not in wb.sheetnames:
        raise XlsxSkillError(f"Sheet not found: {sheet_name}")
    return wb[sheet_name]


def _op_create_sheet(wb: Any, op: dict[str, Any]) -> str:
    name = safe_sheet_name(op.get("sheet") or op.get("name") or "Sheet")
    if name in wb.sheetnames:
        raise XlsxSkillError(f"Sheet already exists: {name}")
    wb.create_sheet(title=name)
    return f"created sheet {name}"


def _op_update_cell(wb: Any, op: dict[str, Any]) -> str:
    ws = _sheet(wb, str(op.get("sheet") or "Sheet1"))
    cell = str(op.get("cell") or "").upper()
    if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]{0,6}", cell):
        raise XlsxSkillError("Invalid cell reference.")
    ws[cell] = op.get("value")
    return f"updated {ws.title}!{cell}"


def _bounded_index(value: Any) -> int:
    try:
        idx = int(value)
    except (TypeError, ValueError) as exc:
        raise XlsxSkillError("Row/column index must be an integer.") from exc
    if idx < 1 or idx > 1_048_576:
        raise XlsxSkillError("Row/column index is outside supported bounds.")
    return idx


def _op_insert_row(wb: Any, op: dict[str, Any]) -> str:
    ws = _sheet(wb, str(op.get("sheet") or "Sheet1"))
    index = _bounded_index(op.get("index") or op.get("row") or 1)
    ws.insert_rows(index, int(op.get("amount") or 1))
    return f"inserted row at {ws.title}!{index}"


def _op_delete_row(wb: Any, op: dict[str, Any]) -> str:
    ws = _sheet(wb, str(op.get("sheet") or "Sheet1"))
    index = _bounded_index(op.get("index") or op.get("row") or 1)
    ws.delete_rows(index, int(op.get("amount") or 1))
    return f"deleted row at {ws.title}!{index}"


def _op_insert_column(wb: Any, op: dict[str, Any]) -> str:
    ws = _sheet(wb, str(op.get("sheet") or "Sheet1"))
    index = _bounded_index(op.get("index") or op.get("column") or 1)
    ws.insert_cols(index, int(op.get("amount") or 1))
    return f"inserted column at {ws.title}!{index}"


def _op_delete_column(wb: Any, op: dict[str, Any]) -> str:
    ws = _sheet(wb, str(op.get("sheet") or "Sheet1"))
    index = _bounded_index(op.get("index") or op.get("column") or 1)
    ws.delete_cols(index, int(op.get("amount") or 1))
    return f"deleted column at {ws.title}!{index}"


def _op_format_range(wb: Any, op: dict[str, Any], openpyxl: Any) -> str:
    ws = _sheet(wb, str(op.get("sheet") or "Sheet1"))
    cell_range = str(op.get("range") or op.get("cell") or "")
    if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]{0,6}(:[A-Z]{1,3}[1-9][0-9]{0,6})?", cell_range.upper()):
        raise XlsxSkillError("Invalid format range.")
    fill = op.get("fill")
    font = op.get("font") or {}
    alignment = op.get("alignment") or {}
    selection = ws[cell_range.upper()]
    rows = selection if isinstance(selection, tuple) else ((selection,),)
    for row in rows:
        cells = row if isinstance(row, tuple) else (row,)
        for cell in cells:
            if fill:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor=str(fill).replace("#", "")[:6])
            if font:
                cell.font = openpyxl.styles.Font(
                    bold=bool(font.get("bold", False)),
                    italic=bool(font.get("italic", False)),
                    color=str(font.get("color", "")).replace("#", "")[:6] or None,
                )
            if alignment:
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal=alignment.get("horizontal"),
                    vertical=alignment.get("vertical"),
                )
    return f"formatted {ws.title}!{cell_range.upper()}"


def _op_add_chart(wb: Any, op: dict[str, Any], openpyxl: Any) -> str:
    ws = _sheet(wb, str(op.get("sheet") or "Sheet1"))
    chart_type = str(op.get("chart_type") or op.get("type") or "bar").lower()
    chart_map = {
        "bar": openpyxl.chart.BarChart,
        "line": openpyxl.chart.LineChart,
        "pie": openpyxl.chart.PieChart,
        "scatter": openpyxl.chart.ScatterChart,
    }
    if chart_type not in chart_map:
        raise XlsxSkillError("Chart type must be one of: line, bar, pie, scatter.")
    data_range = str(op.get("data_range") or "")
    anchor = str(op.get("anchor") or "E2").upper()
    if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]{0,6}:[A-Z]{1,3}[1-9][0-9]{0,6}", data_range.upper()):
        raise XlsxSkillError("Chart data_range must be an Excel range.")
    if not re.fullmatch(r"[A-Z]{1,3}[1-9][0-9]{0,6}", anchor):
        raise XlsxSkillError("Chart anchor must be an Excel cell reference.")

    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(data_range.upper())
    data = openpyxl.chart.Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    chart = chart_map[chart_type]()
    chart.title = str(op.get("title") or f"{chart_type.title()} Chart")
    chart.style = int(op.get("style") or 10)
    if chart_type == "scatter":
        xvalues = openpyxl.chart.Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
        yvalues = openpyxl.chart.Reference(ws, min_col=max_col, min_row=min_row + 1, max_row=max_row)
        series = openpyxl.chart.Series(yvalues, xvalues, title_from_data=False)
        chart.series.append(series)
    else:
        chart.add_data(data, titles_from_data=True)
    chart.height = float(op.get("height") or 7.5)
    chart.width = float(op.get("width") or 12)
    ws.add_chart(chart, anchor)
    return f"added {chart_type} chart to {ws.title}!{anchor}"
